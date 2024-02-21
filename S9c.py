import openpyxl
import os
import time

# Function to read and process the Excel files
def process_excel_file(file_path):
    column1_set = set()
    column4_set = set()

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(values_only=True):
        if len(row) >= 4:
            column1_set.add(row[0])
            column4_set.add(row[3])

    return column1_set, column4_set

# Get the paths of both Excel files from the user
file1_path = input("Enter the path to the first Excel file: ")
file2_path = input("Enter the path to the second Excel file: ")

# Measure the time taken
start_time = time.time()

# Process the first Excel file
file1_column1, file1_column4 = process_excel_file(file1_path)

# Process the second Excel file
file2_column1, file2_column4 = process_excel_file(file2_path)

# Find common values in column 1 and 4
common_column1 = file1_column1 & file2_column1
common_column4 = file1_column4 & file2_column4

# Create a new workbook and sheet to save the matched data
result_wb = openpyxl.Workbook()
result_sheet = result_wb.active

# Write the matched data to the new sheet
for value1, value4 in zip(common_column1, common_column4):
    result_sheet.append([value1, value4])

# Save the result to a new file
result_path = os.path.join("C:\\python\\ChatGPTclass", "same.xlsx")
result_wb.save(result_path)

# Calculate the time taken
end_time = time.time()
elapsed_time = end_time - start_time

print(f"Matched data saved to '{result_path}'")
print(f"Time taken: {elapsed_time:.2f} seconds")
